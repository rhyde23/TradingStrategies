#Excel Sheet Testing

#Imported Libraries

#openpyxl library - For opening and manipulating excel sheets
import openpyxl

#datetime library - To get current week and day
from datetime import datetime, timedelta

#The "calculate_stats" function calculates the variety of metrics used to evaluate a trading strategy's performance like Profit Factor and Win Percentage
def calculate_stats(exited_trades, max_holding, wins=0, failed_trades=0, gains=0, losses=0, max_holding_past=0) :

    #If the "max_holding" input accessed from the all-time stats for this strategy is not an integer or float, set it to zero.
    if type(max_holding) not in [int, float] :
        max_holding = 0

    #If the "wins" input accessed from the all-time stats for this strategy is not an integer or float, set it to zero.
    if type(wins) not in [int, float] :
        wins = 0
        
    #If the "failed_trades" input accessed from the all-time stats for this strategy is not an integer or float, set it to zero.
    if type(failed_trades) not in [int, float] :
        failed_trades = 0
        
    #If the "gains" input accessed from the all-time stats for this strategy is not an integer or float, set it to zero.
    if type(gains) not in [int, float] :
        gains = 0
        
    #If the "losses" input accessed from the all-time stats for this strategy is not an integer or float, set it to zero.
    if type(losses) not in [int, float] :
        losses = 0
        
    #If the "max_holding_past" input accessed from the all-time stats for this strategy is not an integer or float, set it to zero.
    if type(max_holding_past) not in [int, float] :
        max_holding_past = 0

    #If there are no exited trades, return N/A for all stats
    if len(exited_trades) == 0 :
        return "N/A", "N/A"

    #Iterate through each exited trade
    for exited_trade in exited_trades :

        #If the stock was bought
        if exited_trade[2] :

            #If the exit price is greater than the entrance price, add one win and add a price-adjusted value to "gains" in order to eventually calculate average win for profit factor
            if exited_trade[3] >= exited_trade[1] :
                wins += 1
                gains += ((exited_trade[3]-exited_trade[1])/exited_trade[1])*100

            #If the exit price is less than the entrance price, add one loss and add a price-adjusted value to "loss" in order to eventually calculate average loss for profit factor
            else :
                losses += ((exited_trade[1]-exited_trade[3])/exited_trade[1])*100
                failed_trades += 1

        #If the stock was sold
        else :

            #If the exit price is less than the entrance price, add one win and add a price-adjusted value to "gains" in order to eventually calculate average win for profit factor
            if exited_trade[3] <= exited_trade[1] :
                wins += 1
                gains += ((exited_trade[1]-exited_trade[3])/exited_trade[1])*100

            #If the exit price is greater than the entrance price, add one loss and add a price-adjusted value to "loss" in order to eventually calculate average loss for profit factor
            else :
                losses += ((exited_trade[3]-exited_trade[1])/exited_trade[1])*100
                failed_trades += 1
   
    #Calculate win percentage by dividing total wins by the number of exited trades
    win_percentage = (wins/(wins+failed_trades))

    #If there are any losses
    if win_percentage != 1 :

        #Calculate "avg_win" and "avg_loss" by dividing gains by the number of wins and losses by the number of losses
        avg_win, avg_loss = gains/wins, losses/failed_trades
   
        #Calculate Profit Factor by using the profit factor formula
        profit_factor = round((win_percentage*avg_win)/((1-win_percentage)*avg_loss), 4)

    #If there were no losses
    else :
        profit_factor = "ALL WINS"

    #Return final stats
    return [profit_factor, round(win_percentage, 4), max(max_holding, max_holding_past), gains, losses, wins, failed_trades]

#The "get_all_time_stats" function gets the all-time stats of each strategy in the ALL TIME sheet
def get_all_time_stats(all_time_sheet) :

    #The "all_time_stats" dictionary puts the all-time stats as the value for each key (strategy name)
    all_time_stats = {}

    #Iterate through each non-header row to populate "all_time_stats"
    for all_time_row_index in range(2, all_time_sheet.max_row+1) :

        #Get the stat list by iterating through each relevant column on this strategy's row
        stat_list = [all_time_sheet.cell(row = all_time_row_index, column = all_time_column_index).value for all_time_column_index in range(2, all_time_sheet.max_column+1)]

        #Assign this stat list to this row's strategy name key
        all_time_stats[all_time_sheet.cell(row = all_time_row_index, column = 1).value] = stat_list

    #Return the "all_time_stats" dictionary
    return all_time_stats

#The "get_current_week_range" function gets the current Sunday-Saturday week that today's date is in for Excel Sheet recording purposes
def get_current_week_range(dt) :

    #Get the current weekday
    weekday = dt.weekday()

    #If the current day is a sunday, set the "start_sunday" to the current day
    if weekday == 6 :
        start_sunday = dt

    #If the current day is not a sunday, set the "start_sunday" to the most recent previous sunday
    else :
        start_sunday = dt-timedelta(weekday+1)

    #Calculate the end saturday to six days after the start sunday
    end_saturday = start_sunday+timedelta(6)

    #Return the string from the sunday --> the saturday for the excel sheet
    return start_sunday.strftime('%Y-%m-%d')+" --> "+end_saturday.strftime('%Y-%m-%d')

#The "write_headers" function writes all the headers at the top of a new sheet
def write_headers(sheet) :

    #The "headers" list contains the names of all the desired headers to be put at the top of the sheet.
    headers = ["NAME", "PROFIT FACTOR", "WIN %", "MAX CONCURRENT", "P-A GAINS", "P-A LOSSES", "# OF WINS", "# OF LOSSES", "DATE"]

    #Iterate through each header inn the list and write it in its correct cell.
    for header_ind, header in enumerate(headers) :
        sheet.cell(row = 1, column = header_ind+1).value = header

#The "record_performance_data" function records performance data for each finished strategy
def record_performance_data(path, finished_strategies) :

    #Load the workbook
    workbook_loaded = openpyxl.load_workbook(path)

    #Get current time
    dt = datetime.now()

    #Try to load the ALL TIME sheet, if it exists
    try :

        #Redirect the sheet to the ALL TIME sheet
        sheet = workbook_loaded["ALL TIME"]

        #Call the "get_all_time_stats" function to get the all-time stats for each strategy.
        all_time_stats = get_all_time_stats(sheet)
       
    #If it doesn't exist, create it
    except :
       
        #Call the "create_sheet" function from openpyxl
        workbook_loaded.create_sheet(title="ALL TIME")

        #Redirect the sheet to this new sheet
        sheet = workbook_loaded["ALL TIME"]

        #Set the "all_time_stats" dictionary as an empty dictionary
        all_time_stats = {}

        #Call the "write_headers" function to populate the ALL TIME sheet with headers
        write_headers(sheet)

        #Save the workbook
        workbook_loaded.save(path)

    #The "new_added" list will contain all the strategies that have not yet been recorded in the ALL TIME sheet
    new_added = []

    #Call the "get_current_week_range" to get the current sheet to record each strategy performance data
    current_week_range = get_current_week_range(dt)

    #Try to load the sheet for the current week
    try :
        sheet = workbook_loaded[current_week_range]

    #If it does not exist, create the sheet for the current week
    except :

        #Call the "create_sheet" function from openpyxl
        workbook_loaded.create_sheet(title=current_week_range)

        #Now access the sheet for this current week
        sheet = workbook_loaded[current_week_range]

        #Call the "write_headers" function to write the headers at the top of this sheet
        write_headers(sheet)

        #Save the workbook
        workbook_loaded.save(path)

    #The "row_ind" integer is the next empty row to record strategy performance data
    row_ind = sheet.max_row+1

    #Iterate through each finished strategy
    for finished_strategy in finished_strategies :

        #Unpack exited trades, max holding, and the name for this finished strategy
        exited_trades, max_holding, strategy_name = finished_strategy[1:]

        #Call the "calculate_stats" to calculate strategy performance metrics
        calculated_stats = calculate_stats(exited_trades, max_holding)

        #Record Name
        sheet.cell(row = row_ind, column = 1).value = strategy_name

        #Iterate through each calculated stat and record it in the sheet
        current_column = 2
        for calculated_stat in calculated_stats :
            sheet.cell(row = row_ind, column = current_column).value = calculated_stat
            current_column += 1

        #Record today's date
        sheet.cell(row = row_ind, column = current_column).value = dt.strftime('%Y-%m-%d')

        #Add one to the row index
        row_ind += 1

        #If this stratey already exists within the ALL TIME sheet
        if strategy_name in all_time_stats :

            #Get the relevant all time stats
            all_time_max_holding, all_time_gains, all_time_losses, all_time_wins, all_time_failed_trades = all_time_stats[strategy_name][2:][:-1]

            #Calculate the new all time stats given the new stats from this day
            all_time_stats[strategy_name] = calculate_stats(exited_trades, max_holding, wins=all_time_wins, failed_trades=all_time_failed_trades, gains=all_time_gains, losses=all_time_losses, max_holding_past=all_time_max_holding)

        #If this strategy doesn't already exist within the ALL TIME sheet
        else :

            #Set this strategy's all time stats as today's stats
            all_time_stats[strategy_name] = calculated_stats

            #Add this strategy to the "new_added" list
            new_added.append(strategy_name)

    #Redirect sheet to the ALL TIME sheet
    sheet = workbook_loaded["ALL TIME"]

    #THIS LOOP WILL EDIT THE EXISTING STRATEGIES ALL TIME STATS
   
    #Iterate through each row in the ALL TIME sheet
    for all_time_row_index in range(2, sheet.max_row+1) :

        #Get the name of the strategy
        all_time_strat_name = sheet.cell(row = all_time_row_index, column = 1).value
       
        #Iterate through each column and update each all time stat for this strategy
        for all_time_stat_index, all_time_stat in enumerate(all_time_stats[all_time_strat_name]) :
            sheet.cell(row = all_time_row_index, column = all_time_stat_index+2).value = all_time_stat

    #THIS LOOP WILL ADD THE NEWLY ADDED STRATEGIES ALL TIME STATS

    #The "current row" index signifies which row is being added to
    current_row = sheet.max_row+1

    #Iterate through each newly added strategy in the all time stats
    for newly_added in new_added :

        #Record the strategy's name in column 1
        sheet.cell(row = current_row, column = 1).value = newly_added

        #Iterate through each of this newly added strategy's stats and add them to the correct column
        for newly_added_stat_index, newly_added_stat in enumerate(all_time_stats[newly_added]) :
            sheet.cell(row = current_row, column = newly_added_stat_index+2).value = newly_added_stat

        #Add the most recent execution date of this strategy (today's date)
        sheet.cell(row = current_row, column = newly_added_stat_index+3).value = dt.strftime('%Y-%m-%d')

        #Add one to the "current_row" index
        current_row += 1
   
    #Save the workbook
    workbook_loaded.save(path)

###TESTING PURPOSES###

"""
finished_strategies = [[{}, [('GOOGL', 150.01, True, 160.01), ('AAPL', 182.52, False, 150), ('MSFT', 410.02, True, 409.02), ("nah", 100, True, 101)], 4, 'AwesomeStrategy']]
record_performance_data("C:/Users/regin/OneDrive/Desktop/TestBook.xlsx", finished_strategies)
"""
